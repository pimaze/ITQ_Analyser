\
import os
import logging
import json
import io # Added for in-memory file handling
import pandas as pd
import openpyxl # Added for Excel generation
from flask import Flask, render_template, request, send_file, session, redirect, url_for # Added send_file, session, redirect, url_for
from werkzeug.utils import secure_filename
import tempfile # Added for secure temporary file handling

app = Flask(__name__)
# IMPORTANT: Set a secret key for session management.
# Replace 'your_very_secret_key_here' with a real, strong secret key.
# You can generate one using os.urandom(24)
app.secret_key = os.environ.get('FLASK_SECRET_KEY', 'your_very_secret_key_here') # Use environment variable or a default

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Constants
VALID_STATUSES = ['Approved', 'Cancelled']
UPDATE_REQUIRED_STATUSES = ['In Review', 'On Hold', 'Ready for Approval', 'Ready for Review', 'Update Required']
ALLOWED_EXTENSIONS = {'csv'}
MAX_FILE_SIZE = 15 * 1024 * 1024 # 15 MB

# --- Load Configuration ---
def load_json_config(filepath, default_value):
    """Loads configuration from a JSON file with error handling."""
    try:
        with open(filepath, 'r') as f:
            return json.load(f)
    except FileNotFoundError:
        logging.warning(f"{filepath} not found. Using default configuration.")
        return default_value
    except json.JSONDecodeError:
        logging.error(f"Error decoding JSON from {filepath}. Using default configuration.")
        return default_value
    except Exception as e:
        logging.error(f"Unexpected error loading {filepath}: {e}. Using default configuration.")
        return default_value

ANALYSIS_CRITERIA = load_json_config('analysis_criteria.json', {
    "missing_sla": {"condition": "df['SLA'].isnull()", "message": "Deliverables with missing SLA"},
    # Add other default criteria here if needed
})

FIELDS_CONFIG = load_json_config('fields_config.json', {
    "mandatory_fields": ['Deliverable Name', 'Time Spent', 'Deliverable Reviewer'],
    "optional_fields": ['Location of Approved Deliverable']
})
MANDATORY_FIELDS = FIELDS_CONFIG.get('mandatory_fields', [])
OPTIONAL_FIELDS = FIELDS_CONFIG.get('optional_fields', [])

# --- Helper Functions ---
def allowed_file(filename):
    """Checks if the file extension is allowed."""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def preprocess_data(df, pod_filter=None, bto_filter=None):
    """Preprocesses the DataFrame: applies filter, removes empty records, converts dates."""
    original_count = len(df)
    logging.info(f"Preprocessing data. Initial row count: {original_count}")

    if pod_filter:
        try:
            df = df[df['Related Project::POD'] == pod_filter]
            logging.info(f"Applied POD filter '{pod_filter}'. Row count: {len(df)}")
        except KeyError:
            raise ValueError("Column 'Related Project::POD' not found in CSV for filtering.")

    if bto_filter:
        try:
            df = df[df['Related Project::BTO'].isin(bto_filter)]
            logging.info(f"Applied BTO filter '{bto_filter}'. Row count: {len(df)}")
        except KeyError:
            raise ValueError("Column 'Related Project::BTO' not found in CSV for filtering.")

    # Define essential columns that should exist
    required_columns = {'Deliverable Name', 'Date Deliverable is Received for Review', 'Deliverable Approval Date'}
    if not required_columns.issubset(df.columns):
        missing = required_columns - set(df.columns)
        raise ValueError(f"Missing required columns in CSV: {', '.join(missing)}")

    # Drop rows where both essential identifiers are missing
    df = df.dropna(subset=['Deliverable Name', 'Date Deliverable is Received for Review'], how='all')
    logging.info(f"Dropped rows with missing essential identifiers. Row count: {len(df)}")

    # Convert date columns safely
    df['Date Deliverable is Received for Review'] = pd.to_datetime(df['Date Deliverable is Received for Review'], errors='coerce')
    df['Deliverable Approval Date'] = pd.to_datetime(df['Deliverable Approval Date'], errors='coerce')
    logging.info("Converted date columns.")

    if len(df) == 0 and original_count > 0:
         logging.warning("No data remaining after applying filters.")
         # Return empty DataFrame instead of raising error immediately
         # The analyze function will handle the empty case later

    return df

def apply_analysis_criteria(df):
    """Applies the analysis criteria and returns a dictionary of results."""
    results = {}
    if df.empty:
        logging.warning("DataFrame is empty, skipping analysis criteria application.")
        return results # Return empty results if no data

    for analysis_key, details in ANALYSIS_CRITERIA.items():
        try:
            # Ensure required columns exist before evaluation
            required_cols_for_eval = [col for col in df.columns if col in details["condition"]] # Crude check, might need refinement
            if not all(col in df.columns for col in required_cols_for_eval):
                 logging.warning(f"Skipping criteria '{analysis_key}' due to missing columns in DataFrame.")
                 continue

            condition = eval(details["condition"], {'df': df, 'pd': pd}) # Pass pandas as pd
            filtered_df = df.loc[condition] # Use .loc for boolean indexing
            deliverable_names = filtered_df['Deliverable Name'].tolist()
            results[analysis_key] = {"message": details["message"], "deliverables": deliverable_names}
        except SyntaxError as e:
            logging.error(f"Syntax error in condition for '{analysis_key}': {details['condition']}. Error: {e}")
            results[analysis_key] = {"message": f"Error evaluating condition: {details['message']}", "deliverables": ["Syntax Error in Criteria"]}
        except KeyError as e:
             logging.error(f"Missing column '{e}' required for criteria '{analysis_key}': {details['condition']}")
             results[analysis_key] = {"message": details["message"], "deliverables": [f"Missing Column: {e}"]}
        except Exception as e:
            logging.exception(f"Error evaluating condition for '{analysis_key}'. Error: {e}") # Log full traceback
            results[analysis_key] = {"message": f"Error evaluating condition: {details['message']}", "deliverables": ["Evaluation Error"]}
    return results

def check_missing_fields(df, fields, message):
    """Checks for missing fields in the DataFrame and returns a dictionary of results."""
    missing_fields_results = {}
    if df.empty:
        logging.warning(f"DataFrame is empty, skipping missing field check for: {message}")
        return {"message": message, "deliverables": missing_fields_results}

    # Ensure all fields to check actually exist in the DataFrame
    fields_to_check = [f for f in fields if f in df.columns]
    missing_cols = set(fields) - set(fields_to_check)
    if missing_cols:
        logging.warning(f"Columns missing for '{message}' check: {', '.join(missing_cols)}")

    if not fields_to_check: # If no valid columns to check
        logging.warning(f"No valid columns found for missing field check: {message}")
        return {"message": message, "deliverables": missing_fields_results}

    try:
        # Check for NaNs or Nulls in the relevant subset of columns
        missing_fields_df = df[df[fields_to_check].isnull().any(axis=1)]

        for index, row in missing_fields_df.iterrows():
            # Check only the valid fields that were found in the dataframe
            missing = [field for field in fields_to_check if pd.isnull(row[field])]
            deliverable_name = row.get('Deliverable Name', f"Row Index {index}") # Fallback if 'Deliverable Name' is missing
            if deliverable_name and missing: # Ensure deliverable name exists and there are missing fields
                missing_fields_results[deliverable_name] = missing
    except Exception as e:
        logging.exception(f"Error during missing field check for '{message}'. Error: {e}")
        # Optionally return an error indicator in the results
        # return {"message": message, "deliverables": {"Error": "An error occurred during check"}}

    return {"message": message, "deliverables": missing_fields_results}


def analyze_deliverables(filepath, pod_filter=None, bto_filter=None):
    """Analyzes a CSV file, returns results or error dictionary."""
    try:
        # Use low_memory=False for potentially mixed-type columns
        df = pd.read_csv(filepath, low_memory=False)
        logging.info(f"Successfully read CSV: {filepath}")
    except FileNotFoundError:
        logging.error(f"File not found: {filepath}")
        return {"error": "Uploaded file not found during analysis."}
    except pd.errors.EmptyDataError:
        logging.warning(f"CSV file is empty: {filepath}")
        return {"error": "The uploaded CSV file is empty."}
    except pd.errors.ParserError as e:
        logging.error(f"Error parsing CSV file: {filepath}. Error: {e}")
        return {"error": "Invalid CSV file format. Please check the file structure."}
    except Exception as e:
        logging.exception(f"Unexpected error reading CSV: {filepath}")
        return {"error": "An unexpected error occurred while reading the CSV file."}

    try:
        df_processed = preprocess_data(df.copy(), pod_filter, bto_filter) # Work on a copy
    except ValueError as e:
        logging.error(f"Error during preprocessing: {e}")
        return {"error": str(e)}
    except Exception as e:
        logging.exception("Unexpected error during preprocessing.")
        return {"error": "An unexpected error occurred during data preprocessing."}

    if df_processed.empty:
        logging.warning("DataFrame is empty after preprocessing. No analysis performed.")
        # Return an empty result set instead of an error, allows showing "no issues" on results page
        return {
             "info": "No matching data found after applying filters.",
             "missing_mandatory_fields": {"message": "Deliverables with missing mandatory fields", "deliverables": {}},
             "missing_optional_fields": {"message": "Deliverables with missing optional fields", "deliverables": {}}
             # Include other keys with empty deliverables if needed by the template
         }


    # Apply core analysis
    results = apply_analysis_criteria(df_processed)

    # Check missing fields
    results["missing_mandatory_fields"] = check_missing_fields(df_processed, MANDATORY_FIELDS, "Deliverables with missing mandatory fields")
    results["missing_optional_fields"] = check_missing_fields(df_processed, OPTIONAL_FIELDS, "Deliverables with missing optional fields")

    logging.info("Analysis complete.")
    return results

# --- Flask Routes ---
@app.route('/', methods=['GET'])
def index():
    session.pop('analysis_results', None) # Clear previous results when visiting index
    return render_template('index.html')

@app.route('/analyze', methods=['POST'])
def analyze():
    if 'file' not in request.files:
        logging.warning("Analyze request received without file part.")
        # Consider flashing a message instead of returning plain text
        return redirect(url_for('index')) # Redirect back

    file = request.files['file']
    pod_filter = request.form.get('pod_filter', '').strip() or None # Treat empty string as None
    bto_filter = request.form.getlist('bto_filter') or None # Treat empty list as None

    if file.filename == '':
        logging.warning("Analyze request received with empty filename.")
        # Flash message: "No file selected."
        return redirect(url_for('index'))

    if not allowed_file(file.filename):
        logging.warning(f"Invalid file type uploaded: {file.filename}")
        # Flash message: "Invalid file type. Only CSV files are allowed."
        return redirect(url_for('index')) # Redirect is better UX

    # Securely save the uploaded file to a temporary location
    try:
        # Check file size *before* saving
        file.seek(0, os.SEEK_END)
        file_length = file.tell()
        if file_length > MAX_FILE_SIZE:
             logging.warning(f"File size exceeds limit: {file_length} bytes.")
             # Flash message: f"File size exceeds the limit of {MAX_FILE_SIZE / (1024 * 1024)} MB."
             return redirect(url_for('index'))
        file.seek(0) # Reset pointer

        # Use tempfile for better security and cleanup
        with tempfile.NamedTemporaryFile(delete=False, suffix=secure_filename(file.filename)) as temp_file:
            file.save(temp_file.name)
            temp_file_path = temp_file.name
        logging.info(f"File saved temporarily to: {temp_file_path}")

        # Run analysis
        analysis_results = analyze_deliverables(temp_file_path, pod_filter, bto_filter)

    except Exception as e:
        # Catch broad exceptions during file saving or analysis setup
        logging.exception("Error during file saving or analysis setup.")
        analysis_results = {"error": f"An unexpected error occurred: {str(e)}"}
    finally:
        # Ensure temporary file is removed
        if 'temp_file_path' in locals() and os.path.exists(temp_file_path):
            try:
                os.remove(temp_file_path)
                logging.info(f"Removed temporary file: {temp_file_path}")
            except OSError as e:
                logging.error(f"Error removing temporary file {temp_file_path}: {e}")

    # Store results in session before rendering
    session['analysis_results'] = analysis_results

    # Check if analysis returned an error
    if 'error' in analysis_results:
        logging.error(f"Analysis failed with error: {analysis_results['error']}")
        # Render results page but display the error prominently
        return render_template('results.html', analysis_results=analysis_results, error_message=analysis_results['error'])
    elif 'warning' in analysis_results:
         logging.warning(f"Analysis completed with warning: {analysis_results['warning']}")
         return render_template('results.html', analysis_results=analysis_results, warning_message=analysis_results['warning'])
    elif 'info' in analysis_results:
         logging.info(f"Analysis info: {analysis_results['info']}")
         # Render results page but maybe show the info message
         return render_template('results.html', analysis_results=analysis_results, info_message=analysis_results['info'])


    logging.info("Analysis successful, rendering results.")
    return render_template('results.html', analysis_results=analysis_results)


@app.route('/export', methods=['GET'])
def export_results():
    analysis_results = session.get('analysis_results')

    if not analysis_results or 'error' in analysis_results or 'info' in analysis_results:
         logging.warning("Export requested but no valid analysis results found in session.")
         # Redirect to index or show an error page/message
         # Flashing a message is a good option here: flash("No results to export.")
         return redirect(url_for('index'))

    try:
        # Create workbook and remove default sheet
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        logging.info("Generating Excel file...")

        # Iterate through each analysis category
        for key, result_data in analysis_results.items():
            # Skip non-dict items or items without 'deliverables'
            if not isinstance(result_data, dict) or 'deliverables' not in result_data:
                continue

            deliverables = result_data['deliverables']
            sheet_name = key.replace('_', ' ').title()[:31] # Max 31 chars for sheet names, make readable
            ws = wb.create_sheet(title=sheet_name)

            # Write Header Row
            headers = ["Deliverable Name"]
            is_missing_fields_check = isinstance(deliverables, dict) # Check if it's the missing fields structure
            if is_missing_fields_check:
                headers.append("Missing Fields")
            ws.append(headers)

            # Write Data Rows
            if is_missing_fields_check:
                # Handle dictionary structure (missing fields)
                 if not deliverables: # Check if the dictionary is empty
                    ws.append(["No issues found in this category", ""]) # Indicate no issues
                 else:
                    for deliverable_name, missing_fields in deliverables.items():
                        ws.append([deliverable_name, ", ".join(missing_fields)])
            else:
                # Handle list structure (standard checks)
                 if not deliverables: # Check if the list is empty
                     ws.append(["No issues found in this category"]) # Indicate no issues
                 else:
                    for deliverable_name in deliverables:
                        ws.append([deliverable_name])

            # Optional: Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter # Get the column name
                for cell in col:
                    try: # Necessary to avoid error on non-string types
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width


        # Save to in-memory stream
        excel_stream = io.BytesIO()
        wb.save(excel_stream)
        excel_stream.seek(0) # Rewind the stream to the beginning

        logging.info("Excel file generated successfully.")

        return send_file(
            excel_stream,
            as_attachment=True,
            download_name='analysis_results.xlsx', # Changed to .xlsx
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        logging.exception("Error generating Excel file.")
        # Redirect or show an error message
        # flash("An error occurred while generating the Excel file.")
        return redirect(url_for('index'))


# --- Main Execution ---
if __name__ == "__main__":
    # Use Gunicorn or another production server instead of app.run for deployment
    # For development:
     app.run(host='0.0.0.0', port=int(os.environ.get("PORT", 5000)), debug=False) # Turn debug off for production testing
